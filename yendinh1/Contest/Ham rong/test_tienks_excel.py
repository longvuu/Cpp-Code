#!/usr/bin/env python3
"""
test_tienks_excel.py

Generates N tests, compiles each TIENKS C++ file found under the root, runs each executable
on every test, collects detailed results and writes to an Excel (.xlsx) file if openpyxl is installed,
otherwise writes a CSV which Excel can open.

Usage:
    python test_tienks_excel.py [root] --tests 100 --max-n 200 --out results.xlsx

Outputs a workbook/CSV with per-test per-executable rows: test_id, input_size, expected, got, pass, exit_code, runtime_s, stderr
"""
import os
import sys
import re
import argparse
import subprocess
import tempfile
import random
import time
import csv
from collections import OrderedDict

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    HAVE_X = True
except Exception:
    HAVE_X = False

CPP_NAME_RE = re.compile(r'(?i)^ti(en)?ks.*\.cpp$')

def find_cpp_files(root):
    res = []
    for dirpath, dirs, files in os.walk(root):
        for f in files:
            if CPP_NAME_RE.match(f):
                res.append(os.path.join(dirpath, f))
    return sorted(res)

import shutil

def compile_cpp(src, out_dir):
    base = os.path.splitext(os.path.basename(src))[0]
    exe = os.path.join(out_dir, base + '.exe')
    compilers = []
    for name in ('g++', 'clang++'):
        p = shutil.which(name)
        if p:
            compilers.append(p)
    common_paths = [r'D:\MinGW\bin\g++.exe', r'C:\MinGW\bin\g++.exe', r'C:\msys64\mingw64\bin\g++.exe', r'C:\msys64\mingw32\bin\g++.exe']
    for p in common_paths:
        if os.path.exists(p) and p not in compilers:
            compilers.append(p)
    if not compilers:
        compilers = ['g++', 'clang++']
    last_log = ''
    for comp in compilers:
        cmds_to_try = [
            [comp, '-std=c++17', '-O2', src, '-o', exe],
            [comp, '-std=c++17', '-O2', src, '-o', exe, '-lstdc++fs'],
        ]
        for cmd in cmds_to_try:
            try:
                proc = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=30)
                log = proc.stdout.decode(errors='replace') + proc.stderr.decode(errors='replace')
                last_log = log
                if proc.returncode == 0:
                    return True, exe, log
            except FileNotFoundError as e:
                last_log = str(e)
            except Exception as e:
                last_log = str(e)
    return False, exe, last_log

# Reference solver
import bisect

def reference_solver_from_text(text):
    lines = [l for l in text.splitlines() if l.strip()]
    if not lines:
        return ''
    n = int(lines[0].strip())
    intervals = []
    for i in range(1, min(len(lines), n+1)):
        l, r, w = map(int, lines[i].split())
        intervals.append((l, r, w))
    intervals.sort(key=lambda x: x[1])
    ends = [r for (_, r, _) in intervals]
    dp = [0] * (n + 1)
    for i in range(1, n+1):
        l, r, w = intervals[i-1]
        k = bisect.bisect_right(ends, l, 0, i-1)
        dp[i] = max(dp[i-1], dp[k] + w)
    return str(dp[n])

def gen_random_test(n, coord_max=1000, weight_max=1000):
    arr = []
    for _ in range(n):
        a = random.randint(0, coord_max)
        b = random.randint(0, coord_max)
        if a == b:
            b = a + 1
        if a > b:
            a, b = b, a
        w = random.randint(0, weight_max)
        arr.append((a, b, w))
    lines = [str(n)] + ['{} {} {}'.format(a, b, w) for (a, b, w) in arr]
    return '\n'.join(lines) + '\n'

def run_exe(exe, input_text, timeout=2):
    try:
        t0 = time.perf_counter()
        proc = subprocess.run([exe], input=input_text.encode(), stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=timeout)
        t1 = time.perf_counter()
        out = proc.stdout.decode(errors='replace').strip()
        err = proc.stderr.decode(errors='replace').strip()
        return proc.returncode, out, err, t1-t0
    except subprocess.TimeoutExpired:
        return None, '', 'TIMEOUT', timeout
    except Exception as e:
        return -1, '', str(e), 0.0

def write_results_xlsx(path, rows, headers):
    if HAVE_X:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'results'
        ws.append(headers)
        for r in rows:
            ws.append(r)
        # auto column width
        for i,col in enumerate(ws.columns, start=1):
            maxlen = 0
            for cell in col:
                v = cell.value
                if v is None:
                    l = 0
                else:
                    l = len(str(v))
                if l > maxlen:
                    maxlen = l
            ws.column_dimensions[get_column_letter(i)].width = min(50, maxlen+2)
        wb.save(path)
    else:
        # fallback to csv
        csv_path = os.path.splitext(path)[0] + '.csv'
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(headers)
            for r in rows:
                w.writerow(r)
        print('openpyxl not installed; wrote CSV to', csv_path)

def main(argv):
    parser = argparse.ArgumentParser()
    parser.add_argument('root', nargs='?', default='.', help='root to search')
    parser.add_argument('--tests', type=int, default=100, help='number of tests to generate')
    parser.add_argument('--max-n', type=int, default=200, help='max intervals per test')
    parser.add_argument('--out', default='tienks_results.xlsx', help='output xlsx path')
    parser.add_argument('--seed', type=int, default=123456, help='random seed')
    args = parser.parse_args(argv)

    random.seed(args.seed)
    cpp_files = find_cpp_files(args.root)
    if not cpp_files:
        print('No TIENKS cpp files found under', args.root)
        return 1

    tmpdir = tempfile.mkdtemp(prefix='tienks_excel_')
    print('Compiling into', tmpdir)
    compiled = OrderedDict()
    for f in cpp_files:
        ok, exe, log = compile_cpp(f, tmpdir)
        compiled[f] = {'ok': ok, 'exe': exe, 'log': log}
        print(f, '->', 'OK' if ok else 'FAILED')

    # prepare tests
    tests = []
    # include some deterministic
    tests.append(('sample-1', '4\n1 4 5\n1 3 8\n3 5 4\n4 6 9\n'))
    tests.append(('single-interval', '1\n0 10 100\n'))
    tests.append(('disjoint', '3\n1 2 5\n3 4 6\n5 6 7\n'))
    tests.append(('nested', '3\n1 10 100\n2 3 10\n4 5 20\n'))
    tests.append(('same-ends', '3\n1 3 5\n1 3 10\n3 5 7\n'))

    # random
    for i in range(max(0, args.tests - len(tests))):
        n = random.randint(0, min(100, args.max_n))
        tests.append((f'rand-{i}', gen_random_test(n, coord_max=1000, weight_max=1000)))

    rows = []
    headers = ['test_id','test_index','n','expected','file','ok','exit_code','got','stderr','runtime_s']

    for ti,(tname, txt) in enumerate(tests):
        expected = reference_solver_from_text(txt)
        nlines = int(txt.splitlines()[0]) if txt.splitlines() else 0
        for f,info in compiled.items():
            if not info['ok']:
                fname = os.path.basename(f)
                rows.append([tname, ti, nlines, expected, fname, 'compile_failed', None, '', info['log'][:200], None])
                continue
            code, out, err, runtime = run_exe(info['exe'], txt, timeout=3)
            # extract last integer as output
            got_tok = ''
            m = re.findall(r'-?\d+', out)
            if m:
                got_tok = m[-1]
            ok = (got_tok == re.search(r'-?\d+', expected).group(0) if re.search(r'-?\d+', expected) else expected == got_tok)
            fname = os.path.basename(f)
            rows.append([tname, ti, nlines, expected, fname, ok, code, got_tok, err, runtime])

    out_path = os.path.abspath(args.out)
    if HAVE_X:
        write_results_xlsx(out_path, rows, headers)
        print('Wrote results to', out_path)
    else:
        write_results_xlsx(out_path, rows, headers)
    return 0

if __name__ == '__main__':
    sys.exit(main(sys.argv[1:]))
